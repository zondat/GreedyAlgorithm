from util import *
from openpyxl.utils import get_column_letter, column_index_from_string
from model import *  
from greedyalgo import *

# Read excel and create input data for the algorithm
data_file = 'data.xlsx'
workbook = open_workbook(data_file)
worksheet = open_sheet(workbook, 'Sheet1')

# Get cities data
cities = []
distance_table_idx = (80, 'A')
distance_table_size = 20
# idx = 0
for col in range(column_index_from_string(distance_table_idx[1]) + 1, column_index_from_string(distance_table_idx[1]) + 1 + distance_table_size):
    # idx += 1
    city_name = get_cell_data(worksheet, distance_table_idx[0], get_column_letter(col))
    city_col = search(cities, city_name)
    if city_col == None:
        city_col = City(city_name)
        cities.append(city_col)
        
    # for row in range(distance_table_idx[0] + 1 + idx, distance_table_idx[0] + 1 + distance_table_size):
    for row in range(distance_table_idx[0] + 1, distance_table_idx[0] + 1 + distance_table_size):
        #print(get_cell_data(worksheet, row, get_column_letter(col)))
        city_name = get_cell_data(worksheet, row, distance_table_idx[1])
        city_row = search(cities, city_name)
        
        if city_row == None:
            city_row = City(city_name)
            cities.append(city_row)
            
        if city_col != city_row:
            distance = get_cell_data(worksheet, row, get_column_letter(col))
            city_col.add_distance(city_row, distance)

# Get Warehouse data
warehouses = []
products = []
warehouses_table_idx = (1, 'A')
nb_warehouse_props = 6
nb_warehouses = 4
nb_products = 3

for i in range(0, nb_products):
    product = Product(get_cell_data(worksheet, warehouses_table_idx[0] + 1, get_column_letter(column_index_from_string(warehouses_table_idx[1]) + 1 + i)))
    # print(product.name)
    products.append(product)

# Warehouse Center in Ninh Binh
row = warehouses_table_idx[0] + 2
city_name = 'Ninh Binh'
city = search(cities, city_name)
dc = Warehouse(city)
for j in range(0, nb_products):
    product = products[j]
    capacity_idx = 1
    capacity = get_cell_data(worksheet, row, get_column_letter(column_index_from_string(warehouses_table_idx[1]) + capacity_idx + j))
    dc.add_capacity(product, capacity)
    
    holding_cost_idx = 10
    holding_cost = get_cell_data(worksheet, row, get_column_letter(column_index_from_string(warehouses_table_idx[1]) + holding_cost_idx + j))
    dc.add_holding_cost(product, holding_cost)

    service_level_idx = 16
    service_level = get_cell_data(worksheet, row, get_column_letter(column_index_from_string(warehouses_table_idx[1]) + service_level_idx + j))
    dc.add_service_level(product, service_level)
    
# add dc to warehouse list
warehouses.append(dc)
        
# Other warehouses
for i in range(0, nb_warehouses):
    row = warehouses_table_idx[0] + 3 + i
    city_name = get_cell_data(worksheet, row, warehouses_table_idx[1])
    city = search(cities, city_name)
    warehouse = Warehouse(city)
    #print(warehouse.city.name)
    
    # add attributes
    for j in range(0, nb_products):
        product = products[j]
        capacity_idx = 1
        capacity = get_cell_data(worksheet, row, get_column_letter(column_index_from_string(warehouses_table_idx[1]) + capacity_idx + j))
        #print(capacity)
        warehouse.add_capacity(product, capacity)
        
        demand_idx = 4
        demand = get_cell_data(worksheet, row, get_column_letter(column_index_from_string(warehouses_table_idx[1]) + demand_idx + j))
        #print(demand)
        warehouse.add_demand(product, demand)
        
        ordering_cost_idx = 7
        ordering_cost = get_cell_data(worksheet, row, get_column_letter(column_index_from_string(warehouses_table_idx[1]) + ordering_cost_idx + j))
        #print(ordering_cost)
        warehouse.add_ordering_cost(product, ordering_cost)

        holding_cost_idx = 10
        holding_cost = get_cell_data(worksheet, row, get_column_letter(column_index_from_string(warehouses_table_idx[1]) + holding_cost_idx + j))
        #print(holding_cost)
        warehouse.add_holding_cost(product, holding_cost)
        
        shortage_cost_idx = 13
        shortage_cost = get_cell_data(worksheet, row, get_column_letter(column_index_from_string(warehouses_table_idx[1]) + shortage_cost_idx + j))
        #print(shortage_cost)
        warehouse.add_shortage_cost(product, shortage_cost)

        service_level_idx = 16
        service_level = get_cell_data(worksheet, row, get_column_letter(column_index_from_string(warehouses_table_idx[1]) + service_level_idx + j))
        #print(service_level)
        warehouse.add_service_level(product, service_level)
    
    # add warehouse to list
    warehouses.append(warehouse)

# Get retailer and business data
business_data = BusinessData()
retailers = []
retailer_table_idx = (9, 'A')
period_row = retailer_table_idx[0] + 1

for city_idx in range(0, len(cities)):
    city_row = retailer_table_idx[0]-1+(city_idx+1)*nb_products
    # print('City row is ' + str(city_row))
    city_name = get_cell_data(worksheet, city_row, retailer_table_idx[1])
    # print('City name is ' + city_name)
    city = search(cities, city_name)
    retailer = Retailer(city)
    
    for product_idx in range(0, nb_products):
        product_row = city_row + 3 - (nb_products - product_idx)
        product_name = get_cell_data(worksheet, product_row, get_column_letter(column_index_from_string(retailer_table_idx[1]) + 1))
        # print('Search product: ' + product_name)
        product = search(products, product_name)
        
        retailer.add_product(product)
        
        for period_idx in range(0, 12):
            period_col = get_column_letter(column_index_from_string(retailer_table_idx[1]) + period_idx + 2)
            period = get_cell_data(worksheet, period_row, period_col)
            value = get_cell_data(worksheet, product_row, period_col)
            # print('Retailer: ' + retailer.city.name)
            # print('Period: ' + str(period))
            # print('Product: ' + product.name)
            # print('Demand: ' + str(value))
            business_data.add_record(period, product, retailer, value)
            
    retailers.append(retailer)
    
# print ('Number of retailer: ' + str(len(retailers)))

# Notations
# m -> warehouse = [1..M]
# j -> retailer  = [1..J]
# p -> product   = [1..P]
# t -> time period  = [1..T]
# C -> Capacity
#   Cp -> Capacity of DC
#   C1 -> Capacity of regional warehouse ==> C1_mp: Capacity of regional warehouse m and for product partition
# O -> Ordering Cost
# H -> Holding Cost of DC
#   H1_mp -> Holding Cost of regional warehouse m for product partition
# b_jp -> shortage cost for retailer j and product partition
# L -> Lead time
# X_mjpt -> Expected demand of product p of retailer j at period t at warehouse m
def C(p):
    return dc.capacity(p)
    
def C1(m, p):
    return warehouses[m].capacity(p)
    
def O(p):
    return dc.ordering_cost(p)
    
def H(p):
    return dc.holding_cost(p)
    
def H1(m, p):
    return warehouses[m].capacity(p)
    
# def b(j, p):
    # return retailers[j]
    
# def L(p):
    # return
    
A = 100000000

# def Q(p, t):
    # return business_data
    
def Q1(m, p, t):
    record = business_data.search_record(t, p.name, warehouses[m].name)
    return record.value
    
# def Q2(m, j, p, t):
    # record = business_data.search_record(t, p.name, warehouses[m].name)
    # return record.value
    
# def a(p, t):
    # if dc.

# Critera and Objective Function
# def constraint_1(p, t):
    # return Q(p, t) <= A*a(p, t)
    
# def constraint_2(p, t):
    # return 1 - Q(p, t) <= A*(1 - a(p, t))

# Objective function
def MinTC():
    sum_1 = 0
    for t in range(1, 13):
        for m in range(1, nb_warehouses):
            for p in range(0, nb_products):
                sum_1 += H1(m, p) * I1(m, p, t)
            
    sum_2 = 0
    for t in range(1, 13):
        for p in range(0, nb_products):
            sum_2 += H(p) * I(p, t)
        
    sum_3 = 0
    for t in range(1, 13):
        for p in range(0, nb_products):
            sum_3 += a(p, t) * O(p)
    
    return sum_1 + sum_2 + sum_3

# Apply Greedy Algorithm to the problem of new warehouse position
new_warehouse_radius = 500
new_warehouse_capacity = 500 

def get_retailers_in_range(city, radius):
    cities_in_range = []
    # print ('City ' + city.name)
    for i in range(0, len(cities)):
        city_dest = cities[i]
        # print ('Destination city ' + cities[i].name)
        if city != city_dest:
            if city.distance_to(city_dest) <= radius:
                cities_in_range.append(city_dest)
    # print('Number of cities in range: ' + str(len(cities_in_range)))
            
    retailers_in_range = []
    for i in range(0, len(cities_in_range)):
        # print('City ' + cities_in_range[i].name)
        for j in range(0, len(retailers)):
            city_of_retailer = retailers[j].city
            # print('Retailer\'s city: ' + city_of_retailer.name)
            if cities_in_range[i].name == city_of_retailer.name:
                retailers_in_range.append(retailers[j])
    # print('Number of retailers in range: ' + str(len(retailers_in_range)))
    
    return retailers_in_range
    
def get_distance(city, retailers):
    total_distance = 0
    for i in range(0, len(retailers)):
        total_distance += city.distance_to(retailers[i].city)
    return total_distance
        
# Constraint 1: total demand of bound retailer in radius is no more than warehouse's capacity    
def constraint_1(city):
    retailers_in_range = get_retailers_in_range(city, new_warehouse_radius)
    is_valid = True
    for period in range(1, 13):
        for j in range(0, len(products)):
            product_name = products[j].name
            total_demand_in_period_of_product = 0
            for k in range(0, len(retailers_in_range)):
                print('Search demand of product ' + product_name + 'in period ' + str(period) + ' of retailer '+ retailers[k].city.name)
                total_demand_in_period_of_product += business_data.get_demand(period, product_name, retailers[k])
                # print('Total demand for ' + product_name + ' in period ' + str(period) 'is ' + str(total_demand_in_period_of_product))
                is_valid = is_valid and total_demand_in_period_of_product <= new_warehouse_capacity           
    return is_valid
    
# Constraint 2: the new warehouse must be different from existed
def constraint_2(city):
    hanoiIdx = 9
    hcmIdx = 5
    ninhbinhIdx = 3
    return city != cities[ninhbinhIdx] and city != cities[hanoiIdx] and city != cities[hcmIdx]
    
# Objective: distance to all retailers in range is minimum
def get_total_distance_to_retailer(city):
    retailer_in_ranges = get_retailers_in_range(city, new_warehouse_radius)
    return get_distance(city, retailer_in_ranges)

greedyAlgo = GreedyAlgorithm(None, get_total_distance_to_retailer)
greedyAlgo.data = cities
greedyAlgo.add_constraint(constraint_1)
greedyAlgo.add_constraint(constraint_2)
new_warehouse_city = greedyAlgo.search_optimum()
print('City of new warehouse ' + new_warehouse_city.name)
retailers_in_range = get_retailers_in_range(new_warehouse_city, 500)
for i in range(0, len(retailers_in_range)):
    retailer = retailers_in_range[i]
    not_already_has_warehouse = True
    for j in range(0, len(warehouses)):
        not_already_has_warehouse = not_already_has_warehouse and retailer.city != warehouses[j].city
    if not_already_has_warehouse:
        print('Retailer ' + retailer.city.name)


# ninh_binh = cities[3]
# print(ninh_binh.name)
# retailers_in_range = get_retailers_in_range(ninh_binh, 500)
# for i in range(0, len(retailers_in_range)):
    # retailer = retailers_in_range[i]
    # not_already_has_warehouse = True
    # for j in range(0, len(warehouses)):
        # not_already_has_warehouse = not_already_has_warehouse and retailer.city != warehouses[j].city
    # if not_already_has_warehouse:
        # print('Retailer ' + retailer.city.name)
    

    
